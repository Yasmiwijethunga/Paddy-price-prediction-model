import csv
import io
from datetime import datetime
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from dependencies import get_db, get_current_user
from repositories import record_repo
from schemas.paddy_record import PaddyRecordOut

router = APIRouter(prefix="/records", tags=["Records"])


@router.get("/years")
def get_available_years(db: Session = Depends(get_db), _=Depends(get_current_user)):
    """Return a sorted list of years that have at least one paddy record."""
    return {"years": record_repo.get_available_years(db)}


@router.get("/", response_model=List[PaddyRecordOut])
def list_records(
    year: Optional[int] = Query(None, description="Filter records by a specific year"),
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=500),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """
    Retrieve paddy data records.
    Optionally filter by year.  Supports pagination via skip/limit.
    """
    if year is not None:
        return record_repo.get_records_by_year(db, year)
    return record_repo.get_all_records(db, skip=skip, limit=limit)


@router.get("/export/csv")
def export_records_csv(
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Download all paddy records as a CSV file."""
    records = record_repo.get_all_records(db, skip=0, limit=10000)

    columns = [
        "id", "year", "season", "variety", "district",
        "cultivated_area", "total_production", "harvest_quantity",
        "rainfall", "temperature",
        "fertilizer_price", "seed_cost", "pesticide_cost",
        "population", "rice_consumption", "paddy_price",
    ]

    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=columns, extrasaction="ignore")
    writer.writeheader()
    for rec in records:
        writer.writerow({col: getattr(rec, col, "") for col in columns})

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=paddy_records.csv"},
    )


@router.get("/export/excel")
def export_records_excel(
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Download all paddy records as a formatted Excel (.xlsx) file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    records = record_repo.get_all_records(db, skip=0, limit=10000)

    columns = [
        "id", "year", "season", "variety", "district",
        "cultivated_area", "total_production", "harvest_quantity",
        "rainfall", "temperature",
        "fertilizer_price", "seed_cost", "pesticide_cost",
        "population", "rice_consumption", "paddy_price",
    ]
    headers = [
        "ID", "Year", "Season", "Variety", "District",
        "Cultivated (ha)", "Production (MT)", "Harvest (MT)",
        "Rainfall (mm)", "Temp (°C)",
        "Fertilizer (LKR/kg)", "Seed (LKR/kg)", "Pesticide (LKR)",
        "Population", "Rice Consumption (MT)", "Paddy Price (LKR/kg)",
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paddy Records"

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(fill_type="solid", fgColor="1B5E20")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    for ri, rec in enumerate(records, 2):
        for ci, col in enumerate(columns, 1):
            ws.cell(row=ri, column=ci, value=getattr(rec, col, None))

    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        iter([buf.read()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=paddy_records.xlsx"},
    )


@router.get("/export/summary")
def export_summary(
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Return a JSON summary report built from live DB data."""
    from repositories import prediction_repo
    records = record_repo.get_all_records(db, skip=0, limit=10000)
    predictions = prediction_repo.get_all_predictions(db)

    price_recs = [r for r in records if r.paddy_price]
    years = sorted({r.year for r in records})

    return {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "scope": {"district": "Anuradhapura", "season": "Maha", "variety": "Nadu"},
        "data_records": len(records),
        "years_covered": years,
        "predictions_run": len(predictions),
        "price_summary": {
            "min":  min(r.paddy_price for r in price_recs) if price_recs else None,
            "max":  max(r.paddy_price for r in price_recs) if price_recs else None,
            "avg":  round(sum(r.paddy_price for r in price_recs) / len(price_recs), 2) if price_recs else None,
            "latest": sorted(price_recs, key=lambda r: r.year)[-1].paddy_price if price_recs else None,
        },
        "production_summary": {
            "avg_cultivated_ha": round(
                sum(r.cultivated_area for r in records if r.cultivated_area) /
                max(len([r for r in records if r.cultivated_area]), 1), 0),
            "avg_production_mt": round(
                sum(r.total_production for r in records if r.total_production) /
                max(len([r for r in records if r.total_production]), 1), 0),
        },
    }


@router.get("/{record_id}", response_model=PaddyRecordOut)
def get_record(
    record_id: int,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    """Retrieve a single paddy record by its primary key."""
    record = record_repo.get_record_by_id(db, record_id)
    if not record:
        raise HTTPException(status_code=404, detail=f"Record id={record_id} not found.")
    return record
